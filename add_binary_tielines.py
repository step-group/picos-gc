# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl"]
# ///
"""Merge the Water-solvent binary tie-line onto each ternary block sheet, in place.

Each `Bloque X` sheet gets a single binary edge tie-line (HBA + HBD + Water, no 2PE)
appended just below its 20-vial ternary grid, at rows 25-28: two phases (solvent-rich
"Superior", water-rich "Inferior"), two replicates each. Raw-measurement cells (masses
D/F/H, volume E, GC areas M/N, KF water U/V) are left BLANK as hand-fill placeholders;
the dilution -> real -> average -> normalized-point formulas are pre-wired and reference
the sheet's own HBA/HBD slopes ($G$2/$H$2). The normalized endpoints land in AE/AF/AG
(HBA/HBD/Water, sum to 1) at row 25 (organic) and row 27 (aqueous).

fill_ternarios.py only touches rows 5-24, so this hand-filled block is invisible to it.
Any old separate `Bloque X bin` sheets (earlier design) are deleted.

Run: uv run add_binary_tielines.py   (edits Sistemas ternarios_MF.xlsx in place; SKIPS
     any sheet whose tie-line block already exists, so re-running never clobbers data).
"""

from __future__ import annotations

import openpyxl

from fill_ternarios import WB_IN, recalc_workbook

_R0, _R1 = 25, 28  # binary tie-line block rows (Superior x2, Inferior x2)

# Per-vial formulas (rows 25-28). Binary = ternary minus the solute column; $G$2/$H$2
# are the sheet's existing HBA/HBD slopes.
_PER_VIAL = {
    "G": "=1000-E{r}",
    "I": "=F{r}-D{r}",
    "J": "=H{r}-F{r}",
    "K": "=J{r}+I{r}",
    "P": "=M{r}/$G$2",  # diluido %m/m = area / slope
    "Q": "=N{r}/$H$2",
    "S": "=P{r}*$K{r}/$I{r}",  # real %m/m = diluido x dilution
    "T": "=Q{r}*$K{r}/$I{r}",
}
# Per-phase formulas on the two anchor rows 25 (Superior) and 27 (Inferior); each phase
# spans rows r, r+1 (2 reps).
_PER_PHASE = {
    "X": "=AVERAGE(S{r}:S{s})/100",  # mean HBA fraction over the two reps
    "Y": "=AVERAGE(T{r}:T{s})/100",  # mean HBD fraction
    "Z": "=AVERAGE(U{r}:V{s})/100",  # KF water
    "AA": "=SUM(W{r}:Z{r})",  # closure (~1); W empty so this is X+Y+Z
    "AE": '=IFERROR(X{r}/$AA{r},"")',  # normalized HBA (blank, not #DIV/0, until filled)
    "AF": '=IFERROR(Y{r}/$AA{r},"")',  # normalized HBD
    "AG": '=IFERROR(Z{r}/$AA{r},"")',  # normalized Water
}


def build_tieline(ws) -> None:
    """Write the binary Water-solvent tie-line block into rows 25-28 of `ws` in place.
    The slopes ($G$2/$H$2), HBA/HBD names (M3/N3) and the AE/AF/AG row-4 headers already
    live on the ternary sheet, so only the block's own cells are written."""
    ws.merge_cells(f"A{_R0}:A{_R1}")
    ws[f"A{_R0}"] = "Bin"  # system code + idempotency sentinel
    for i, r in enumerate(range(_R0, _R1 + 1)):
        ws[f"B{r}"] = "Superior" if i < 2 else "Inferior"
        ws[f"C{r}"] = i + 1  # vial code 1..4
        for col, f in _PER_VIAL.items():
            ws[f"{col}{r}"] = f.format(r=r)
    for r in (_R0, _R0 + 2):  # phase-anchor rows 25 (organic) and 27 (aqueous)
        for col, f in _PER_PHASE.items():
            ws[f"{col}{r}"] = f.format(r=r, s=r + 1)


def add_binary_tielines(path=WB_IN) -> None:
    wb = openpyxl.load_workbook(path, data_only=False)  # keep formulas

    for name in [n for n in wb.sheetnames if n.endswith(" bin")]:
        del wb[name]  # earlier separate-sheet design; superseded by the merged block
        print(f"deleted {name}")

    for name in wb.sheetnames:
        ws = wb[name]
        if ws[f"A{_R0}"].value:  # tie-line block already present -> idempotent
            print(f"skip    {name} (tie-line already present)")
            continue
        build_tieline(ws)
        print(f"tieline {name}")

    wb.save(path)
    # openpyxl drops formula caches on save, so the ternary sheets would open blank until
    # recalc. Restore them (and compute the new tie-line formulas) via LibreOffice.
    warn: list[str] = []
    recalc_workbook(path, warn)
    print(f"\nWrote {path}")
    for m in warn:
        print("  -", m)


def demo() -> None:
    """Check the binary normalization identity: with no solute term, AE+AF+AG =
    (X+Y+Z)/(X+Y+Z) = 1 for any nonzero phase."""
    vial, plus_samp, plus_met = 2.75, 2.95, 3.51  # D, F, H weighings, g
    area_hba, area_hbd = 900.0, 850.0  # M, N GC areas
    kf1, kf2 = 8.5, 8.6  # U, V KF water %m/m, two reps
    slope_hba, slope_hbd = 176.3849, 171.2346
    musa = plus_samp - vial  # I "Muestra, g"
    mtot = (plus_met - plus_samp) + musa  # K "Mue + Met"
    x = area_hba / slope_hba * mtot / musa / 100  # real HBA fraction
    y = area_hbd / slope_hbd * mtot / musa / 100  # real HBD fraction
    z = (kf1 + kf2) / 2 / 100  # KF water
    aa = x + y + z  # closure (solute term empty)
    ae, af, ag = x / aa, y / aa, z / aa
    assert abs(ae + af + ag - 1.0) < 1e-12, (ae, af, ag)
    assert all(0 < v < 1 for v in (ae, af, ag)), (ae, af, ag)
    print("demo OK")


if __name__ == "__main__":
    import sys

    if "--check" in sys.argv:
        demo()
    else:
        add_binary_tielines()
