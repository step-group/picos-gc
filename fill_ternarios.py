# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl"]
# ///
"""Fill `Sistemas ternarios_MF.xlsx` with picos-gc peak areas + emit a results CSV.

Each sheet `Bloque A..I` is a pseudoternary LLE workbook: Water + 2-phenylethanol
(Solute) + a two-terpene solvent (HBA + HBD). We write the GC areas into columns
L/M/N per vial; the sheet's own formulas turn area -> diluted %m/m -> real %m/m
(x dilution factor) -> 2-replicate average -> KF water -> normalized ternary point.

All Pend.CC slopes are sourced from CC_MF.xlsx (through-origin fit over the SM stock
point + E1..E5 — including SM reproduces the workbook's own slopes exactly). This
fills every block, and replaces the stale 363.58 2PhEt value in E/F/H/I with 187.31.

Prereq: `uv run python label_terpenos.py` has produced out/<batch>/merged_samples.csv.
Run:    uv run fill_ternarios.py   (PEP 723 header pulls openpyxl; paths are
        script-relative, so it works from any directory)
        (also shells out to pcsaft-quaternary/plot_experimental_tielines.py at the
        end, so the ternary diagrams under out/tielines/ are regenerated in one go)
"""

from __future__ import annotations

import argparse
import csv
import re
import shutil
import subprocess
import tempfile
from glob import glob
from pathlib import Path

import openpyxl

_ROOT = Path(__file__).resolve().parent  # script-relative so CWD doesn't matter
WB_IN = _ROOT / "Sistemas ternarios_MF.xlsx"
WB_OUT = _ROOT / "Sistemas ternarios_MF_filled.xlsx"
CC_MF = _ROOT / "CC_MF.xlsx"
DATA = _ROOT / "FLECK_TERPENOS2026"
OUT_CSV = _ROOT / "out" / "ternarios_resultados.csv"
BIN_OUT_CSV = _ROOT / "out" / "binarios_tielines.csv"  # computed Water-solvent binary endpoints
BINARIOS_CSV = _ROOT / "out" / "BINARIOS_TERPENOS" / "samples.csv"
# BIN system -> ternary block, SEQUENTIAL (BIN n = the nth block in order A,B,C,D,E,F,H,I):
# 1=ThyCarvone(A) 2=ThyGer(B) 3=ThyCarvacrol(C) 4=ThyEugenol(D) 5=ThyCamphor(E)
# 6=CamphorCarvone(F) 7=CamphorCarvacrol(H) 8=CamphorEugenol(I). Carvone reads as
# "Geraniol" in samples.csv (the binary-method RT is miscalibrated), so blocks A & F route
# their Carvone area through fill_binary's single-remaining-terpene fallback (right value,
# wrong label). All 8 blocks map; a block only resolves an endpoint once its masses + KF
# are hand-entered on rows 25-28, so today only block A computes.
BIN_TO_BLOCK = {1: "A", 2: "B", 3: "C", 4: "D", 5: "E", 6: "F", 7: "H", 8: "I"}

TERPS = ("camph", "carvone", "carvacrol", "geraniol", "thymol", "eugenol")
DISPLAY = {  # canon -> workbook-style name written into blank M3/N3 cells
    "camphor": "Camphor",
    "carvone": "Carvone",
    "carvacrol": "Carvacrol",
    "geraniol": "Geraniol",
    "thymol": "Thymol",
    "eugenol": "Eugenol",
    "2pe": "2PhEt",
}
# sample codes vary by batch: A1T1, B1-T1, E1_T1, even BM1-T1 (block F). Match the
# trailing (system number, phase, rep); prefix letters/separators are ignored.
CODE_RE = re.compile(r"(\d+)[-_ ]?([TB])([12])\s*$", re.I)

# Composition QC. KF water is an independent absolute assay, so anchor it and split
# the GC remainder by ratio rather than normalising all four by their (often <1) sum
# (that rescales the good water — the D1/I1 spurious-curve bug). Σ outside the band ⇒
# mass didn't close; a dominant GC peak whose two replicate injections differ
# >MISMATCH_MAX x (a uniform whole-sample dilution slip) ⇒ replicate_mismatch.
CLOSURE_BAND = (0.5, 1.5)
MISMATCH_MAX = 3.0
MISMATCH_MIN_FRAC = 0.1  # only judge mismatch on a real constituent, not trace wobble
ORGANIC_WATER_MAX = 0.5  # KF anchor only the water-poor (organic) phase; see results_rows


def parse_key(code: str) -> tuple[int, str, int] | None:
    m = CODE_RE.search(str(code).strip())
    return (int(m.group(1)), m.group(2).upper(), int(m.group(3))) if m else None


def canon(s) -> str:
    """Canonical compound key shared across sheet names, CSV headers, CC_MF sheets."""
    k = re.sub(r"[^a-z0-9]", "", str(s).lower())
    for key in TERPS:
        if key in k:
            return "camphor" if key == "camph" else key
    if "phenyl" in k or k in ("2pe", "2phet", "pe"):
        return "2pe"
    return k


def hba_priority(wb) -> dict[str, float]:
    """Per-terpene tendency to be the HBA, learned from already-labelled sheets.

    HBA appearance scores 0, HBD scores 1, averaged; lower = more HBA-like. Used to
    order a block's two terpenes when its M3/N3 names are blank (e.g. Bloque F).
    """
    seen: dict[str, list[int]] = {}
    for ws in wb.worksheets:
        hba, hbd = ws["M3"].value, ws["N3"].value
        if hba and hbd:
            seen.setdefault(canon(hba), []).append(0)
            seen.setdefault(canon(hbd), []).append(1)
    return {k: sum(v) / len(v) for k, v in seen.items()}


def cc_mf_slopes(xlsx: Path) -> dict[str, float]:
    """{canon: through-origin slope (area per %m/m)} from CC_MF.xlsx standards.

    Fit over EVERY (%m/m, Área) pair in cols N/O — the SM stock point plus E1..E5.
    Including SM is what reproduces the workbook's own Pend.CC exactly (excluding it
    is ~1-5% low). slope = Σ(wt·area)/Σ(wt²).
    """
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    out: dict[str, float] = {}
    for ws in wb.worksheets:
        pts = [
            (float(r[13]), float(r[14]))
            for r in ws.iter_rows(values_only=True)
            if len(r) > 14 and isinstance(r[13], int | float) and isinstance(r[14], int | float)
        ]
        if len(pts) >= 2:
            sxy = sum(w * a for w, a in pts)
            sxx = sum(w * w for w, _ in pts)
            out[canon(ws.title)] = sxy / sxx
    return out


def load_vial_areas(merged_csv: Path) -> dict[tuple[int, str, int], dict[str, float]]:
    """{(sysnum, phase, rep): {canon_compound: area_mean}} from a merged CSV."""
    rows = list(csv.reader(merged_csv.open()))
    hdr = rows[0]
    cols = {i: canon(h[:-10]) for i, h in enumerate(hdr) if h.endswith("_area_mean")}
    out: dict[tuple[int, str, int], dict[str, float]] = {}
    for r in rows[1:]:
        if r and r[0] and (key := parse_key(r[0])) is not None:
            out[key] = {c: float(r[i] or 0) for i, c in cols.items()}
    return out


def load_binary_areas(
    csv_path: Path, warn: list[str]
) -> dict[tuple[int, str, int], dict[str, float]]:
    """{(binnum, phase, rep): {canon_compound: area}} from BINARIOS samples.csv.

    Same shape as load_vial_areas, but the columns end in `_area` (per-injection, no
    `_area_mean` averaging). Missing file -> {} + a warning (binary areas just stay blank).
    """
    if not csv_path.exists():
        warn.append(f"binary: {csv_path} missing - run label_terpenos.py (skipping binary fill)")
        return {}
    rows = list(csv.reader(csv_path.open()))
    hdr = rows[0]
    cols = {i: canon(h[:-5]) for i, h in enumerate(hdr) if h.endswith("_area")}
    out: dict[tuple[int, str, int], dict[str, float]] = {}
    for r in rows[1:]:
        if r and r[0] and (key := parse_key(r[0])) is not None:
            out[key] = {c: float(r[i] or 0) for i, c in cols.items()}
    return out


# BIN phase/rep -> tie-line row: Superior (organic) 25/26, Inferior (aqueous) 27/28.
_BIN_ROWS = {("T", 1): 25, ("T", 2): 26, ("B", 1): 27, ("B", 2): 28}


def fill_binary(ws, block: str, bin_areas: dict, warn: list[str]) -> None:
    """Write the Water-solvent tie-line HBA/HBD areas into M/N of rows 25-28.

    Areas come from BINARIOS samples.csv via BIN_TO_BLOCK; only M/N are written (masses
    D/F/H and KF U/V stay hand-entered). Every block maps to a BIN, so the guard below is
    just a safety net. Must run after fill_block so M3/N3 (HBA/HBD names) are set.
    """
    binnum = next((b for b, blk in BIN_TO_BLOCK.items() if blk == block), None)
    if binnum is None:
        return  # unmapped block (shouldn't happen — all 8 map)
    if ws["A25"].value != "Bin":
        warn.append(f"{block}: no binary tie-line block (run add_binary_tielines.py first)")
        return
    hba_key, hbd_key = canon(ws["M3"].value or ""), canon(ws["N3"].value or "")
    for (ph, rep), row in _BIN_ROWS.items():
        a = bin_areas.get((binnum, ph, rep))
        if a is None:
            warn.append(f"{block}: binary BIN{binnum} {ph}{rep} absent from samples.csv")
            continue
        terp = {c: v for c, v in a.items() if c != "ipa" and v}  # nonzero terpene areas
        if hba_key not in terp:
            warn.append(f"{block}: binary HBA {hba_key!r} not found in BIN{binnum} {ph}{rep}")
            continue
        ws[f"M{row}"] = terp[hba_key]
        if hbd_key in terp:
            ws[f"N{row}"] = terp[hbd_key]
        else:  # HBD label differs (classifier) -> use the single remaining terpene
            other = [v for c, v in terp.items() if c != hba_key]
            if len(other) == 1:
                ws[f"N{row}"] = other[0]
                warn.append(
                    f"{block}: binary HBD {hbd_key!r} not labelled in BIN{binnum} {ph}{rep}; "
                    "used the other terpene"
                )
            else:
                warn.append(f"{block}: binary HBD ambiguous in BIN{binnum} {ph}{rep} ({list(terp)})")


def _binary_bydiff(ws, reps: tuple[int, int], g2, h2) -> list[float] | None:
    """Binary endpoint (HBA, HBD, water) by difference — no KF. Per rep: dilution
    df=(H-D)/(F-D), HBA=M/g2*df/100, HBD=N/h2*df/100, water=1-HBA-HBD; average the reps.
    Returns None if slopes, masses, or areas are missing (can't compute without them)."""
    if not (g2 and h2):
        return None
    hbas, hbds = [], []
    for r in reps:
        d, f, h = _num(ws, f"D{r}"), _num(ws, f"F{r}"), _num(ws, f"H{r}")
        m, n = _num(ws, f"M{r}"), _num(ws, f"N{r}")
        if None in (d, f, h, m, n) or (f - d) == 0:
            continue
        df = (h - d) / (f - d)
        hbas.append(m / g2 * df / 100)
        hbds.append(n / h2 * df / 100)
    if not hbas:
        return None
    hba, hbd = sum(hbas) / len(hbas), sum(hbds) / len(hbds)
    return [hba, hbd, 1.0 - hba - hbd]


def binary_tieline_rows(wb_d) -> list[list]:
    """Water-solvent binary endpoints from a recalc'd (data_only) workbook: rows 25/27,
    cols AE/AF/AG = HBA/HBD/water mass fractions (2PE = 0).

    The pre-wired formula is KF-anchored. When its endpoint resolves (areas+masses+KF)
    we use it (`water_src=kf`); when KF is blank (e.g. the operator dropped the unreliable
    high-water aqueous KF) the formula errors, so we fall back to water **by difference**
    from areas+masses (`water_src=bydiff`) — but only for a genuinely aqueous endpoint
    (resulting water > 0.5), since by-difference is unreliable for the water-poor organic
    phase. A block with no masses yields neither and contributes nothing. Skips the
    degenerate pure-water point (AE=AF=0 when masses+KF but no areas are present)."""
    out = []
    for sheet in wb_d.sheetnames:
        ws = wb_d[sheet]
        if ws["A25"].value != "Bin":
            continue
        block = sheet.split()[-1]
        hba, hbd = ws["M3"].value, ws["N3"].value
        g2, h2 = _num(ws, "G2"), _num(ws, "H2")
        for row, reps in ((25, (25, 26)), (27, (27, 28))):
            ae, af, ag = _num(ws, f"AE{row}"), _num(ws, f"AF{row}"), _num(ws, f"AG{row}")
            src = "kf"
            if None in (ae, af, ag) or (ae == 0 and af == 0):
                bd = _binary_bydiff(ws, reps, g2, h2)
                if bd is None or bd[2] <= 0.5:  # no data, or organic phase w/o KF (untrustworthy)
                    continue
                ae, af, ag = bd
                src = "bydiff"
            phase = "aqueous" if ag > 0.5 else "organic"
            out.append([block, phase, hba, _fmt(ae), hbd, _fmt(af), _fmt(ag), src])
    return out


def _num(ws, coord: str) -> float | None:
    v = ws[coord].value
    return float(v) if isinstance(v, int | float) else None


def _vial_rows(ws):
    """Yield (row, sysnum, ph, rep, phase) for the 20 data rows (5..24).

    System number is the digit in `Cód. Sist` (its letter is unreliable: blocks
    F/H/I carry a stale `E1..E5`). Block identity comes from the sheet name.
    """
    sysnum, counts = None, {}
    for row in range(5, 25):
        a = ws[f"A{row}"].value
        if a:
            m = re.search(r"\d+", str(a))
            sysnum, counts = (int(m.group()) if m else sysnum), {}
        phase = str(ws[f"B{row}"].value or "").strip()
        ph = "T" if phase.lower().startswith("sup") else "B"
        counts[ph] = counts.get(ph, 0) + 1
        yield row, sysnum, ph, counts[ph], phase


def fill_block(
    ws, ws_d, block: str, areas: dict, cc: dict, priority: dict, warn: list[str]
) -> list[dict]:
    """Write missing names/slopes + L/M/N areas; return per-vial records for the CSV.

    `ws` is the formula-preserving sheet we write into; `ws_d` is the data_only
    twin, used to read the cached numeric masses/KF (those columns are formulas).
    """
    # Auto-label blank HBA/HBD names from this block's terpenes, ordered by the
    # convention learned from labelled sheets (Bloque F ships unlabelled).
    terps = sorted(
        {k for a in areas.values() for k in a if k != "2pe"},
        key=lambda c: priority.get(c, 1.0),
    )
    auto = []
    if not ws["M3"].value and terps:
        ws["M3"] = DISPLAY.get(terps[0], terps[0])
        auto.append(f"HBA={ws['M3'].value}")
    if not ws["N3"].value:
        hbd = next((t for t in terps if t != canon(ws["M3"].value or "")), None)
        if hbd:
            ws["N3"] = DISPLAY.get(hbd, hbd)
            auto.append(f"HBD={ws['N3'].value}")
    if auto:
        warn.append(
            f"{block}: auto-labelled {', '.join(auto)} from batch terpenes (confirm HBA/HBD roles)"
        )

    hba_key, hbd_key = canon(ws["M3"].value or ""), canon(ws["N3"].value or "")
    if not ws["M3"].value or not ws["N3"].value:
        warn.append(f"{block}: HBA/HBD names still blank; terpene areas not placed")

    # All slopes come from CC_MF (through-origin, SM-inclusive) — the single source of
    # truth. For A-D this re-writes the same values already there; for E/F/H/I it fills
    # the blank terpene slopes and replaces the stale 363.58 2PhEt value with 187.31.
    for cell, comp in (("F2", "2pe"), ("G2", hba_key), ("H2", hbd_key)):
        if comp in cc:
            new = round(cc[comp], 4)
            old = _num(ws, cell)
            if old is not None and abs(old - new) > 0.5:
                warn.append(
                    f"{block}: {cell} {old} -> {new} (CC_MF {comp}; was it a separate calibration?)"
                )
            ws[cell] = new
        elif comp:
            warn.append(f"{block}: no CC_MF slope for {comp!r} ({cell})")

    recs = []
    for row, sysnum, ph, rep, phase in _vial_rows(ws):
        a = areas.get((sysnum, ph, rep))
        if a is None:
            warn.append(
                f"{block}: vial row {row} ({block}{sysnum}{ph}{rep}) absent from merged_samples"
            )
            continue
        if "2pe" in a:
            ws[f"L{row}"] = a["2pe"]
        if hba_key in a:
            ws[f"M{row}"] = a[hba_key]
        if hbd_key in a:
            ws[f"N{row}"] = a[hbd_key]
        # Dilution from the raw weighings — I ("Muestra, g") = F-D, K ("Mue+Met") = H-D —
        # rather than the sheet's cached I/K *formulas*. data_only reads a formula's last
        # cached value, which is stale if a mass is edited without an Excel recalc; the
        # raw D/F/H cells are always live. Verified F-D/H-D == the I/K formulas in every
        # block (160/160 rows). KF (U/V) and areas (L/M/N) are already raw/live.
        d, fg, hg = _num(ws_d, f"D{row}"), _num(ws_d, f"F{row}"), _num(ws_d, f"H{row}")
        recs.append(
            {
                "row": row,
                "sysnum": sysnum,
                "ph": ph,
                "phase": phase,
                "L": a.get("2pe"),
                "M": a.get(hba_key),
                "N": a.get(hbd_key),
                "I": (fg - d) if (fg is not None and d is not None) else None,
                "K": (hg - d) if (hg is not None and d is not None) else None,
                "U": _num(ws_d, f"U{row}"),
                "V": _num(ws_d, f"V{row}"),
            }
        )
    return recs


def _fmt(x: float | None) -> str:
    return f"{x:.5f}" if x is not None else ""


def kf_anchor(sol: float, hba: float, hbd: float, water: float) -> tuple[list[float], float]:
    """KF-anchored ternary point + closure Σ. Trust the KF water absolutely and split
    the remaining (1 - water) among the GC components by their measured ratio. Where
    mass already closes (Σ≈1) this equals proportional normalisation; it diverges only
    by the closure gap — exactly where the GC totals are untrustworthy."""
    closure = sol + hba + hbd + water
    gc = sol + hba + hbd
    w = min(max(water, 0.0), 1.0)
    if gc <= 0:
        return [0.0, 0.0, 0.0, 1.0], closure
    scale = (1.0 - w) / gc
    return [sol * scale, hba * scale, hbd * scale, w], closure


def _replicate_mismatch(g: list[dict], w, x, y) -> float:
    """Max/min replicate area ratio of the phase's dominant GC component, gated to real
    constituents (raw frac > MISMATCH_MIN_FRAC). A uniform whole-sample scale error
    between the two injections (a dilution slip) shows here as a big ratio on the
    dominant peak. 0.0 when not applicable."""
    col, frac = max(
        (("L", w), ("M", x), ("N", y)),
        key=lambda cf: cf[1] if cf[1] is not None else -1.0,
    )
    if frac is None or frac <= MISMATCH_MIN_FRAC:
        return 0.0
    areas = [r[col] for r in g if r.get(col)]
    if len(areas) < 2 or min(areas) <= 0:
        return 0.0
    return max(areas) / min(areas)


def results_rows(ws, block: str, recs: list[dict], aqueous_bydiff: bool = True) -> list[list]:
    """Replicate the sheet formula chain -> one normalized ternary point per (system, phase).

    Organic (water-poor) phases are always KF-anchored. Aqueous (water-rich) phases use
    water BY DIFFERENCE (`aqueous_bydiff=True`, default) or KF proportional normalisation
    (`False`, legacy). The chosen source is emitted per row as `water_src` (kf|bydiff)."""
    f2, g2, h2 = _num(ws, "F2"), _num(ws, "G2"), _num(ws, "H2")
    hba_name, hbd_name = ws["M3"].value, ws["N3"].value
    groups: dict[tuple[int, str], list[dict]] = {}
    for r in recs:
        groups.setdefault((r["sysnum"], r["ph"]), []).append(r)  # one point per system+phase

    out = []
    for (sysnum, ph), g in sorted(groups.items()):
        system = f"{block}{sysnum}"
        phase = "Superior" if ph == "T" else "Inferior"
        # Per-vial composition (fractions) so a replicate that sampled a *different
        # phase* than its pair can be dropped — averaging a mixed pair otherwise makes a
        # mid-triangle phantom (E2 Superior = one genuine aqueous vial + one that read
        # ~96% KF water yet carried a full organic terpene load).
        vials = []
        for r in g:
            df = (r["K"] / r["I"]) if (r["K"] and r["I"]) else None
            if df is None:
                continue
            s = (r["L"] / f2 * df / 100) if (f2 and r["L"] is not None) else None
            a = (r["M"] / g2 * df / 100) if (g2 and r["M"] is not None) else None
            b = (r["N"] / h2 * df / 100) if (h2 and r["N"] is not None) else None
            ks = [c / 100 for c in (r["U"], r["V"]) if c is not None]
            vials.append({"s": s, "a": a, "b": b, "kf": ks, "raw": r})

        def _is_mixup(v):
            # Majority-water AND majority-organic at once: impossible for one phase, so
            # this vial sampled the wrong phase. (Concentrated organics close >1 but are
            # water-poor; aqueous vials are organic-poor — neither trips this.)
            water = sum(v["kf"]) / len(v["kf"]) if v["kf"] else 0.0
            gc = sum(c for c in (v["s"], v["a"], v["b"]) if c is not None)
            return water > 0.5 and gc > 0.5

        keep = [v for v in vials if not _is_mixup(v)]
        use = keep if keep else vials  # if every vial is a mixup, keep all (stays flagged)
        dropped = 0 < len(use) < len(vials)

        def avg(xs):
            return sum(xs) / len(xs) if xs else None

        sol = [v["s"] for v in use if v["s"] is not None]
        hba = [v["a"] for v in use if v["a"] is not None]
        hbd = [v["b"] for v in use if v["b"] is not None]
        kf = [c for v in use for c in v["kf"]]
        w, x, y, z = avg(sol), avg(hba), avg(hbd), avg(kf)
        flags = [] if (g2 and h2) else ["slopes_missing"]
        closure = None
        src = ""
        if None not in (w, x, y, z):
            closure = w + x + y + z
            if z < ORGANIC_WATER_MAX:
                # Water-poor (organic) phase: KF titrates its small water content
                # reliably, so anchor it and split the rest by GC ratio — this fixes the
                # closure-inflated D1/I1 curve. (By-difference is wrong here: water is the
                # minor component, so it would absorb all the GC closure error.)
                norm, _ = kf_anchor(w, x, y, z)
                src = "kf"
            elif aqueous_bydiff:
                # Water-rich (aqueous) phase, water BY DIFFERENCE: water = 1 - Σ(organics).
                # At ~0.95+ water the KF titration is unreliable, and water so dominates
                # that it's insensitive to the method; the dissolved GC species stay as
                # measured. `closure` (below) keeps the raw KF sum as a drift diagnostic.
                norm = [w, x, y, max(0.0, 1.0 - (w + x + y))]
                src = "bydiff"
            else:
                # Water-rich phase, KF-normalised (legacy --aqueous-water kf).
                tot = closure if closure > 0 else 1.0
                norm = [w / tot, x / tot, y / tot, z / tot]
                src = "kf"
            lo, hi = CLOSURE_BAND
            if not (lo <= closure <= hi):
                flags.append("low_closure")
            if _replicate_mismatch([v["raw"] for v in use], w, x, y) > MISMATCH_MAX:
                flags.append("replicate_mismatch")
            if dropped:
                flags.append("dropped_replicate")
        else:
            norm = [None, None, None, None]
            flags.append("incomplete")
        # Write the final point straight into AD..AG as static values so the workbook
        # matches this CSV. The in-sheet formulas average the phase's rows and can't
        # express the per-vial mixup drop (e.g. E2), so they'd otherwise show a phantom.
        if norm[0] is not None:
            anchor = min(r["row"] for r in g)
            for col, val in zip(("AD", "AE", "AF", "AG"), norm):
                ws[f"{col}{anchor}"] = val
        out.append(
            [
                block,
                system,
                phase,
                _fmt(norm[0]),
                hba_name,
                _fmt(norm[1]),
                hbd_name,
                _fmt(norm[2]),
                _fmt(norm[3]),
                _fmt(closure),
                len(use),
                ";".join(flags),
                src,
            ]
        )
    return out


# Result formulas, generated uniformly for every block. Only Bloque A shipped with
# these (and only partially — its AD..AG summary was Superior-only, rows 5-13), so we
# write the full set rather than copy A's gaps. Inputs (L/M/N areas, masses D-K, KF
# U/V, slopes F2-H2, names M3/N3) live in other cells and are never touched.
PER_VIAL = {  # one per data row 5..24
    "O": "=L{r}/$F$2",
    "P": "=M{r}/$G$2",
    "Q": "=N{r}/$H$2",  # diluted %m/m = area/slope
    "R": "=O{r}*$K{r}/$I{r}",
    "S": "=P{r}*$K{r}/$I{r}",
    "T": "=Q{r}*$K{r}/$I{r}",  # x dilution
}
PER_PHASE = {  # on the first row r of each (system, phase) pair; averages rows r, r+1
    "W": "=AVERAGE(R{r}:R{s})/100",
    "X": "=AVERAGE(S{r}:S{s})/100",  # mean fraction over reps
    "Y": "=AVERAGE(T{r}:T{s})/100",
    "Z": "=AVERAGE(U{r}:V{s})/100",  # Z = KF water
    "AA": "=SUM(W{r}:Z{r})",  # raw closure (≈1); flags a bad phase even after a drop
    # AD..AG (the normalized ternary point) are NOT formulas: results_rows writes them
    # as static values so the workbook matches the CSV. In-sheet formulas would average
    # both replicate rows and so can't express the per-vial mixup drop (e.g. E2).
}


def write_formulas(wb) -> None:
    """Write the full result-formula set into every block (per-vial + per-phase)."""
    for ws in wb.worksheets:
        for r in range(5, 25):
            for col, f in PER_VIAL.items():
                ws[f"{col}{r}"] = f.format(r=r)
        for r in range(5, 25, 2):  # phase-pair anchor rows 5,7,...,23
            for col, f in PER_PHASE.items():
                ws[f"{col}{r}"] = f.format(r=r, s=r + 1)


# Forces LibreOffice to recompute formulas on load (OOXMLRecalcMode 0 = always).
RECALC_XCU = """<?xml version="1.0" encoding="UTF-8"?>
<oor:items xmlns:oor="http://openoffice.org/2001/registry" xmlns:xs="http://www.w3.org/2001/XMLSchema" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load"><prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop></item>
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load"><prop oor:name="ODFRecalcMode" oor:op="fuse"><value>0</value></prop></item>
</oor:items>"""


def recalc_workbook(path: Path, warn: list[str]) -> None:
    """Recompute all formulas in place via headless LibreOffice (openpyxl can't)."""
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        warn.append("LibreOffice not found: open the workbook and press Ctrl+Shift+F9 to recalc")
        return
    with tempfile.TemporaryDirectory() as tmp:
        prof = Path(tmp) / "profile"
        (prof / "user").mkdir(parents=True)
        (prof / "user" / "registrymodifications.xcu").write_text(RECALC_XCU)
        out = Path(tmp) / "out"
        out.mkdir()
        r = subprocess.run(
            [
                soffice,
                "--headless",
                f"-env:UserInstallation=file://{prof}",
                "--calc",
                "--convert-to",
                "xlsx",
                "--outdir",
                str(out),
                str(path),
            ],
            capture_output=True,
            text=True,
            timeout=180,
        )
        produced = out / path.name
        if produced.exists():
            shutil.move(str(produced), str(path))
            print(f"Recalculated {path} via LibreOffice")
        else:
            warn.append(
                f"LibreOffice recalc failed ({r.stderr.strip()[:120]}); open + Ctrl+Shift+F9"
            )


def run_tieline_plots(warn: list[str]) -> None:
    """Run the pcsaft-quaternary tie-line plotter on the CSV we just wrote. It lives
    in a separate uv project (its venv has python-ternary/matplotlib, which this one
    does not), so shell out with `uv run` in that dir rather than importing it."""
    plot_dir = Path(__file__).parent / "pcsaft-quaternary"
    if not (plot_dir / "plot_experimental_tielines.py").exists():
        warn.append(f"tie-line plotter not found under {plot_dir}")
        return
    r = subprocess.run(
        ["uv", "run", "python", "plot_experimental_tielines.py"],
        cwd=plot_dir,
        capture_output=True,
        text=True,
    )
    print(r.stdout, end="")
    if r.returncode != 0:
        warn.append(f"tie-line plotting failed: {r.stderr.strip()[:200]}")


def main() -> None:
    ap = argparse.ArgumentParser(description="Fill the ternary workbook + emit results CSVs.")
    ap.add_argument(
        "--aqueous-water",
        choices=["difference", "kf"],
        default="difference",
        help="aqueous-phase water: 'difference' (1-Σorganics, default) or 'kf' (KF-normalised). "
        "Organic phases are always KF-anchored.",
    )
    args = ap.parse_args()
    aqueous_bydiff = args.aqueous_water == "difference"
    print(f"aqueous-phase water: {args.aqueous_water}")

    cc = cc_mf_slopes(CC_MF)
    wb = openpyxl.load_workbook(WB_IN, data_only=False)  # keep formulas (write target)
    wb_d = openpyxl.load_workbook(WB_IN, data_only=True)  # cached values (read masses/KF)
    priority = hba_priority(wb)  # HBA/HBD ordering learned from labelled sheets
    warn: list[str] = []
    all_rows: list[list] = []
    bin_areas = load_binary_areas(BINARIOS_CSV, warn)  # tie-line areas (rows 25-28)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        block = sheet.split()[-1]  # "Bloque A" -> "A"
        folders = glob(str(DATA / f"{block}1T1 AL *"))
        if not folders:
            warn.append(f"{block}: no FLECK data folder")
            continue
        merged = _ROOT / "out" / Path(folders[0]).name / "merged_samples.csv"
        if not merged.exists():
            warn.append(f"{block}: {merged} missing - run label_terpenos.py")
            continue
        areas = load_vial_areas(merged)
        recs = fill_block(ws, wb_d[sheet], block, areas, cc, priority, warn)
        fill_binary(ws, block, bin_areas, warn)  # Water-solvent tie-line M/N (rows 25-28)
        all_rows += results_rows(ws, block, recs, aqueous_bydiff)
        print(
            f"Bloque {block}: filled {len(recs)}/20 vials "
            f"(HBA={ws['M3'].value}, HBD={ws['N3'].value}, 2PhEt slope={_num(ws, 'F2')})"
        )

    write_formulas(wb)
    wb.save(WB_OUT)
    recalc_workbook(WB_OUT, warn)
    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open("w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(
            [
                "block",
                "system",
                "phase",
                "solute_2phet",
                "HBA",
                "HBA_wt",
                "HBD",
                "HBD_wt",
                "water",
                "closure",
                "n_vials",
                "flags",
                "water_src",
            ]
        )
        w.writerows(all_rows)
    # Water-solvent binary tie-line endpoints (rows 25-28) for the plotter. Read from the
    # recalc'd file so the pre-wired AE/AF/AG formulas have cached values.
    wb_bin = openpyxl.load_workbook(WB_OUT, data_only=True)
    bin_rows = binary_tieline_rows(wb_bin)
    with BIN_OUT_CSV.open("w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["block", "phase", "HBA", "HBA_wt", "HBD", "HBD_wt", "water", "water_src"])
        w.writerows(bin_rows)
    print(f"\nWrote {WB_OUT}\nWrote {OUT_CSV}\nWrote {BIN_OUT_CSV} ({len(bin_rows)} binary endpoint(s))")
    run_tieline_plots(warn)
    if warn:
        print("\nWARNINGS:")
        for m in warn:
            print("  -", m)


def _selfcheck() -> None:
    s = cc_mf_slopes(CC_MF)
    # SM-inclusive fit must reproduce the workbook's own Pend.CC slopes exactly.
    for comp, expect in (
        ("2pe", 187.3142),
        ("thymol", 176.3849),
        ("carvone", 171.2346),
        ("geraniol", 166.2694),
    ):
        assert abs(s[comp] - expect) < 0.05, f"{comp}: {s[comp]} != {expect}"
    a = load_vial_areas(_ROOT / "out" / "A1T1 AL A5B2" / "merged_samples.csv")
    assert abs(a[(1, "T", 1)]["2pe"] - 143.81) < 1, a[(1, "T", 1)]
    assert canon("2PE") == "2pe" and canon("DL-Camphor") == "camphor", "canon broken"
    assert parse_key("BM1-T1") == (1, "T", 1) and parse_key("E3_B2") == (3, "B", 2), "parse_key"
    print("selfcheck OK")


if __name__ == "__main__":
    main()
