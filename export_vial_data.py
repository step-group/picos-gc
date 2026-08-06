# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl"]
# ///
"""Export the per-vial LLE inputs the uncertainty analysis needs.

fill_ternarios.py collapses each phase to a 2-vial average and emits only the
final composition. An uncertainty budget needs the level below that: the two
vials separately, their own weighings, and both Karl-Fischer readings. This
writes that level out verbatim -- no averaging, no unit changes, blanks stay
blank -- for data/raw/lle_ternary/ in the thesis repo.

Also exports the GC calibration standards from CC_MF.xlsx. fill_ternarios reads
them only to take a through-origin slope; the uncertainty analysis needs the
points themselves, because the inverse-prediction interval (and the intercept
test that showed the through-origin form is wrong at low concentration) cannot
be recovered from a slope.

Run: uv run export_vial_data.py
"""

from __future__ import annotations

import csv
from pathlib import Path

import openpyxl

_ROOT = Path(__file__).resolve().parent
WB = _ROOT / "Sistemas ternarios_MF_filled.xlsx"
CC = _ROOT / "CC_MF.xlsx"
MERGED = _ROOT / "out"
OUT_VIALS = _ROOT / "out" / "vial_measurements.csv"
OUT_CAL = _ROOT / "out" / "gc_calibration.csv"

# Rows 5-24 are the five ternary systems (4 rows each: 2 Superior, 2 Inferior);
# rows 25-28 are the water-solvent binary endpoint in the same pattern.
TERNARY_ROWS = range(5, 25)
BINARY_ROWS = range(25, 29)


def _num(ws, coord):
    v = ws[coord].value
    return v if isinstance(v, (int, float)) else None


def injections_per_vial() -> dict[str, int]:
    """{block: n_rep} -- injections per vial, from label_terpenos' merged CSVs.

    2 for blocks B, C, E, F and 1 for A, D, H, I. A single injection gets no
    1/M benefit in the inverse-prediction interval, so this difference between
    blocks is real and must not be averaged away downstream.
    """
    out = {}
    for path in sorted(MERGED.glob("*/merged_samples.csv")):
        block = path.parent.name[0]  # "A1T1 AL A5B2" -> "A"
        with path.open() as fh:
            reps = {int(r["n_rep"]) for r in csv.DictReader(fh) if r.get("n_rep")}
        if reps:
            out[block] = max(reps)
    return out


def _rows(ws, block, rows, kind, n_rep):
    out, code = [], None
    for r in rows:
        code = ws[f"A{r}"].value or code  # the code is written once per system
        phase_position = ws[f"B{r}"].value
        if phase_position is None:
            continue
        out.append(
            {
                "block": block,
                "kind": kind,
                "system": str(code).strip(),
                "phase_position": str(phase_position).strip(),
                "vial": _num(ws, f"C{r}"),
                "n_injections": n_rep,
                "m_vial_g": _num(ws, f"D{r}"),
                "m_vial_sample_g": _num(ws, f"F{r}"),
                "m_vial_sample_diluent_g": _num(ws, f"H{r}"),
                "area_2pe": _num(ws, f"L{r}"),
                "area_hba": _num(ws, f"M{r}"),
                "area_hbd": _num(ws, f"N{r}"),
                "kf_water_pct_1": _num(ws, f"U{r}"),
                "kf_water_pct_2": _num(ws, f"V{r}"),
                "hba": ws["M3"].value,
                "hbd": ws["N3"].value,
            }
        )
    return out


def vial_rows():
    wb = openpyxl.load_workbook(WB, data_only=True)
    n_rep = injections_per_vial()
    rows = []
    for sheet in wb.sheetnames:
        block = sheet.replace("Bloque", "").strip()
        ws = wb[sheet]
        m = n_rep.get(block, 1)
        rows += _rows(ws, block, TERNARY_ROWS, "ternary", m)
        rows += _rows(ws, block, BINARY_ROWS, "binary", m)
    return rows


def calibration_rows():
    """Every (%m/m, area) standard in cols N/O -- the SM stock point plus E1..E5.

    Same selection fill_ternarios.cc_mf_slopes fits its through-origin slope over,
    so a slope refitted from this CSV reproduces the one the workbook used.
    """
    wb = openpyxl.load_workbook(CC, data_only=True)
    rows = []
    for ws in wb.worksheets:
        n = 0
        for r in ws.iter_rows(values_only=True):
            if len(r) > 14 and isinstance(r[13], (int, float)) and isinstance(r[14], (int, float)):
                rows.append(
                    {"compound": ws.title, "standard": n, "w_pct_m_m": r[13], "area": r[14]}
                )
                n += 1
    return rows


if __name__ == "__main__":
    OUT_VIALS.parent.mkdir(parents=True, exist_ok=True)
    for path, rows in ((OUT_VIALS, vial_rows()), (OUT_CAL, calibration_rows())):
        with path.open("w", newline="") as fh:
            w = csv.DictWriter(fh, fieldnames=list(rows[0]))
            w.writeheader()
            w.writerows(rows)
        print(f"Wrote {path} ({len(rows)} rows)")
